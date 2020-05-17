from io import BytesIO

from django.db.models.fields import BooleanField, IntegerField, CharField
from django.http import HttpResponse, HttpResponseBadRequest
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime
from openpyxl import Workbook, load_workbook

from cms.forms.UploadFileForm import UploadFileForm
from cms.helpers.ServiceHelper import ServiceHelper
from cms.models import Service
from django.db.models.fields.related import ManyToManyField, ForeignKey



def getForeignObject(ForeignModel, value):
    try:
        if hasattr(ForeignModel, 'slug'):
            return ForeignModel.objects.get(slug=value)
        if hasattr(ForeignModel, 'code'):
            return ForeignModel.objects.get(code=value)
        if hasattr(ForeignModel, 'name'):
            return ForeignModel.objects.get(name=value)
    except ForeignModel.DoesNotExist:
        return None


def row_to_object(Model, headers, data):
    slug = None
    for i, header in enumerate(headers):
        slug = data[i] if header == 'slug' else slug

    if slug is None:
        object = Model()
    elif Model.objects.filter(slug=slug).exists():
        object = Model.objects.get(slug=slug)
    else:
        object = Model(slug=slug)

    for n, attr_name in enumerate(headers):
        column_type = object._meta.get_field(attr_name)
        value = data[n]

        if isinstance(column_type, ForeignKey) and value:
            model = getForeignObject(column_type.related_model, value)
            if model:
                object.__setattr__(attr_name + '_id', model.id)
        elif isinstance(column_type, ManyToManyField) and value:
            object.save()
            for v in value.split(','):
                model = getForeignObject(column_type.related_model, v)
                if model:
                  object.__getattribute__(attr_name).add(model)
        elif isinstance(column_type, IntegerField) and value:
            object.__setattr__(attr_name, int(value))
        elif isinstance(column_type, BooleanField):
            object.__setattr__(attr_name, True if value.lower() == 'yes' else False)
        elif isinstance(column_type, CharField):
            object.__setattr__(attr_name, value if value else '')
        elif value:
            object.__setattr__(attr_name, value)



    object.save()
    for n, attr_name in enumerate(headers):
        column_type = object._meta.get_field(attr_name)
        value = data[n]

        if isinstance(column_type, ManyToManyField) and value:
            for item in value.split(','):
                model = getForeignObject(column_type.related_model, item)
                if model:
                    getattr(object, attr_name).add(model)

    object.save()


@csrf_exempt
def import_services(request):
    form = UploadFileForm(request.POST, request.FILES)
    if form.is_valid() == False:
        return HttpResponseBadRequest('The file you uploaded is invalid')

    file_in_memory = request.FILES['file'].read()
    wb = load_workbook(filename=BytesIO(file_in_memory))
    ws = wb.active

    headers = []
    for n, row in enumerate(ws.iter_rows()):
        data = []
        for cell in row:
            headers.append(cell.value) if n == 0 else data.append(cell.value)

        if len(data):
            model = row_to_object(Service, headers, data)
        try:
            pass
        except Exception as e:
            raise e
            return HttpResponseBadRequest(str(e))

    return HttpResponse('File imported successfully')


def export_services(request):
    '''
    Download file by category
    http://localhost:8000/admin/export_services/?category=1
    '''
    cat = int(request.GET['category'])
    service = Service()
    helper = ServiceHelper(service)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-services.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Services'

    # define available columns
    columns = ['slug', 'category', 'meta_description']
    rows = []
    for field in helper.fields:
        if cat in field.categories:
            columns.append(field.key)

    services = Service.objects.filter(category=cat)
    for service in services:
        row = []
        rows.append(row)
        helper = ServiceHelper(service)
        for column in columns:
            row.append(helper.get_csv_value(column))

    # add columns to excel's first row
    for num, title in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=num)
        cell.value = title

    for rnumber, row in enumerate(rows, 2):
        for cnumber, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=rnumber, column=cnumber)
            cell.value = str(cell_value)

    workbook.save(response)
    return response
